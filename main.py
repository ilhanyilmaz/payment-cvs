import os
import sys
import argparse

from datetime import timedelta

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

def boolean_input(bool_question):
    bool_input = input(bool_question + " (y/n)\n")
    if bool_input != "y" and bool_input != "n":
        return boolean_input(bool_question)
    elif bool_input == 'y':
        return True
    else:
        return False

def range_input(integer_question, min_value, max_value):
    integer_input = input(integer_question + "("+ str(min_value) + "-" + str(max_value) +")")
    try:
        integer_input = int(integer_input)
    except ValueError:
        print("Try again!")
        return range_input(integer_question, min_value, max_value)
    if integer_input<min_value or integer_input>max_value:
        print("Try again!")
        return range_input(integer_question, min_value, max_value)
    
    return integer_input

def choice_input(choice_question, choices):
    print(choice_question)
    choice_options = []
    for choice in choices:
        choice_options.append(choice[0])
        print("["+choice[0]+"] : " + choice[1])

    print("[x] : exit")
    selected_input = input("Your choice: ")
    choice_options.append("x")
    if selected_input in choice_options:
        return selected_input
    else:
        print("Try again!")
        return choice_input(choice_question, choices)


parser = argparse.ArgumentParser(description='Create csv for calendar from payment xls.')
parser.add_argument('-f', '--filename', type=str, help='Worksheet filename')
parser.add_argument('-o', '--output_filename', type=str, help='Output filename')
parser.add_argument('-ow', '--overwrite', action='store_true', help='Overwrite output file')
parser.add_argument('-e', '--extend', action='store_true', help='Extend output file')
parser.add_argument('-hr', '--hour', type=int, help="Start Time")
parser.add_argument('-d', '--duration', type=int, help="Duration in hours")
parser.add_argument('-r', '--reminder', action='store_true', help="Reminder On/Off")
parser.add_argument('-rm', '--remindminute', type=int, action='append', nargs='+', help="Remind x minutes before")
parser.add_argument('-rh', '--remindhour', type=int, action='append', nargs='+',  help="Remind x hours before")
parser.add_argument('-rd', '--remindday', type=int, action='append', nargs='+',  help="Remind x days before")
args = parser.parse_args()

if not args.filename:
    print("Please provide a file with option -f")
    sys.exit()

wb = None
ws = None
try:
    wb = load_workbook(filename= args.filename, read_only=True, data_only=True)
    ws = wb.active
    print(ws.title)
except InvalidFileException:
    print("File not valid")
    sys.exit()

filename = args.filename
output_filename = filename[:filename.rfind(".")-1] + '.csv'

csv_file = None

if os.path.exists(output_filename):
    if args.extend:
        csv_file = open(output_filename, 'a')
    elif not args.overwrite:
        print("csv file already exists")
        cont = boolean_input("Would you like to overwrite it?")
        if not cont:
            sys.exit()

if csv_file == None:
    csv_file = open(output_filename, 'w')

input_time = 0
if args.hour:
    if args.hour < 0 or args.hour > 23:
        print('--hour parameter shout be between 0-23')
        sys.exit()
    input_time = args.hour
else:
    input_time = range_input("At what time of the day you want to add", 0, 23)

duration = 1

if args.duration:
    duration = args.duration

location = "."
if args.duration:
    location = args.location

#########################################
### Parsing reminders

add_reminder = False
if args.reminder:
    add_reminder = args.reminder
else:
    add_reminder = boolean_input("Would you like to set up a reminder?")

reminders = []
if add_reminder:
    if not args.remindminute and not args.remindhour and not args.remindday:
        while True:
            user_input = choice_input("Please select one", [['m','x minutes before'],['h', 'x hours before'],['d', 'x days before']])
            if user_input == 'x':
                break
            elif user_input == 'm':
                reminder_minute = range_input("How many minutes before?", 0, 60)
                reminders.append(['m', reminder_minute])
            elif user_input == 'h':
                reminder_hour = range_input("How many hours before?", 0, 24)
                reminders.append(['h', reminder_hour])
            elif user_input == 'd':
                reminder_day = range_input("How many days before?", 0, 30)
                reminders.append(['d', reminder_day])
    else:
        if args.remindminute:
            for option in args.remindminute[0]:
                reminders.append(['m', option])
        if args.remindhour:
            for option in args.remindhour[0]:
                reminders.append(['h', option])
        if args.remindday:
            for option in args.remindday[0]:
                reminders.append(['d', option])

### End of Parsing reminders
#########################################
### Start of Adding file parameters

if not args.extend:
    csv_file.write("Subject,Start Date,Start Time,End Date, End Time, Description,Location,Reminder On/Off")
    
    for reminder in reminders:
        csv_file.write(",Reminder Date,Reminder Time")

    csv_file.write("\n")

### End of Adding parameters
#########################################
### Start of Adding entries

def db_num(value):
    if value<10:
        return "0"+str(value)
    return str(value)

def getDateStr(date):
    return db_num(date.day) + "." + db_num(date.month) + "." + str(date.year) + "," + db_num(date.hour) + ":" + db_num(date.minute)

def getHourStr(hour):
    return str(hour)+":00"

for i in range(1,ws.max_row+1):
    cRow=ws[i]
    event_date = cRow[0].value + timedelta(hours=input_time)
    end_date = event_date + timedelta(hours=duration)
    csv_file.write(cRow[1].value+",")
    csv_file.write(getDateStr(event_date)+",")
    csv_file.write(getDateStr(end_date) + ",")
    csv_file.write(cRow[2].value+",")
    csv_file.write(location+",")

    if add_reminder:
        csv_file.write("TRUE")
        reminder_datetimes = []
        for reminder in reminders:
            csv_file.write(",")
            reminder_datetime = None
            if reminder[0] == 'm':
                reminder_datetime = event_date - timedelta(minutes=reminder[1])
            elif reminder[0] == 'h':
                reminder_datetime = event_date - timedelta(hours=reminder[1])
            elif reminder[0] == 'd':
                reminder_datetime = event_date - timedelta(days=reminder[1])

            csv_file.write(getDateStr(reminder_datetime))
    else:
        csv_file.write("FALSE")
    csv_file.write("\n")
csv_file.close()
    